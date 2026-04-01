import openpyxl
import os
from pipeline.engine.extractors.extractor_factory import ExtractorFactory

def create_dummy_excel(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add noise
    ws.cell(row=1, column=1, value="NCRB REPORT 2023")
    ws.cell(row=2, column=1, value="Note: This is a dummy file.")
    
    # Add Headers at Row 4
    headers = ["SL", "State/UT", "District", "Total Cognizable Crimes", "Ignore This"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    
    # Add Data
    data = [
        [1, "Maharashtra", "Mumbai", 5000, "foo"],
        [2, "Maharashtra", "Pune", 3000, "bar"],
        [None, "TOTAL", None, 8000, "baz"]
    ]
    for r_idx, row in enumerate(data, 5):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
            
    wb.save(path)
    return path

def test_phase3():
    print("Starting Phase 3 Verification...")
    test_xl = "temp_test_data.xlsx"
    create_dummy_excel(test_xl)
    
    try:
        # Mock YAML Config for this test
        config = {
            'header_detection': {
                'method': 'pattern_match',
                'pattern': 'SL'
            },
            'column_mapping': {
                'State/UT': {'field': 'state'},
                'District': {'field': 'district'},
                'Total Cognizable Crimes': {'field': 'total_cases'}
            },
            'row_filters': {
                'summary_patterns': ['TOTAL']
            }
        }
        
        extractor = ExtractorFactory.get_extractor('xlsx', config)
        results = extractor.extract(test_xl)
        
        print(f"\nExtracted {len(results)} records.")
        for r in results:
            print(f"   Record: {r}")
            
        # Assertions
        assert len(results) == 2, f"Expected 2 data records, got {len(results)}"
        assert results[0]['state'] == "Maharashtra"
        assert results[0]['district'] == "Mumbai"
        assert results[0]['total_cases'] == 5000
        
        # Check if TOTAL was filtered out correctly
        for r in results:
            if r['state'] == "TOTAL":
                print("❌ Error: Summary row 'TOTAL' was not filtered!")
                return False
        
        print("\nPHASE 3 VERIFIED: Pattern-based extraction and column mapping are working with 100% accuracy.")
        return True

    except Exception as e:
        print(f"ERROR: Phase 3 verification failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        if os.path.exists(test_xl):
            os.remove(test_xl)

if __name__ == "__main__":
    success = test_phase3()
    exit(0 if success else 1)
