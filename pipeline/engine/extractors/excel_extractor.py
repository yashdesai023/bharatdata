import openpyxl
import re
from pipeline.engine.extractors.base_extractor import BaseExtractor

class ExcelExtractor(BaseExtractor):
    def extract(self, file_path):
        """Extracts data and identifies the summary row."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        records = []
        summary_row = None
        header_row_idx = None
        
        # 1. Detect Header Row
        header_config = self.config.get('header_detection', {})
        method = header_config.get('method')
        
        if method == 'fixed_row':
            header_row_idx = header_config.get('row', 1)
        elif method == 'pattern_match':
            pattern = header_config.get('pattern')
            # Support a 'patterns' list in config
            patterns = header_config.get('patterns')
            if patterns and isinstance(patterns, list):
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    for p in patterns:
                        if any(re.search(p, str(cell).strip(), re.IGNORECASE) for cell in row if cell):
                            header_row_idx = row_idx
                            break
                    if header_row_idx:
                        break
            else:
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if pattern and any(re.search(pattern, str(cell).strip(), re.IGNORECASE) for cell in row if cell):
                        header_row_idx = row_idx
                        break

            # Fallback: try multi-row header detection by combining consecutive rows
            if not header_row_idx:
                rows = list(sheet.iter_rows(values_only=True))
                nrows = len(rows)
                # Try start positions within first 10 rows
                for start in range(0, min(10, nrows)):
                    if not any(rows[start]):
                        continue
                    # Try combining 2 or 3 consecutive rows
                    for span in (2, 3):
                        if start + span > nrows:
                            continue
                        combined = []
                        maxcols = max(len(r) for r in rows[start:start+span])
                        for col in range(maxcols):
                            parts = []
                            for r in range(start, start+span):
                                cell = rows[r][col] if col < len(rows[r]) else None
                                if cell and str(cell).strip():
                                    parts.append(str(cell).strip())
                            combined.append(" ".join(parts).strip() if parts else None)

                        # Heuristic: if combined headers include several known tokens, accept
                        tokens = ['state','district','village','name','tot_p','total','population','area']
                        matched = sum(1 for c in combined if c and any(t in c.lower() for t in tokens))
                        if matched >= 3:
                            header_row_idx = start + span
                            # Store combined headers to use directly
                            raw_headers = combined
                            break
                    if header_row_idx:
                        break

        # If header row not found yet, raise
        if not header_row_idx and 'raw_headers' not in locals():
            raise ValueError(f"Could not find header row in {file_path}")

        # 2. Map Columns
        raw_headers = [cell for cell in next(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))]
        col_map = self._map_columns(raw_headers)
        
        # 3. Process Data
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not any(row): continue
            
            # Identify Summary Row (e.g. Total)
            if self._is_summary_row(row):
                summary_data = {}
                for col_idx, val in enumerate(row):
                    header = raw_headers[col_idx]
                    field = col_map.get(header)
                    if field: summary_data[field] = val
                summary_row = summary_data
                continue # Do not include summary in records
            
            # Skip Rows
            if self._should_skip_row(row):
                continue
            
            record = {}
            for col_idx, val in enumerate(row):
                header = raw_headers[col_idx]
                field = col_map.get(header)
                if field: record[field] = val
            
            if record:
                records.append(record)
                
        return {
            'records': records,
            'summary': summary_row,
            'col_map': col_map
        }
